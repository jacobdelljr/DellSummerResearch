from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from PIL import Image, UnidentifiedImageError
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = PROJECT_ROOT / "data" / "raw" / "HiRISE Data"
SPLIT_WORKBOOK = PROJECT_ROOT / "data" / "splits" / "HiRise data - cross-validation - 7 versions (version 1).xlsx"
METRICS_DIR = PROJECT_ROOT / "outputs" / "metrics"
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png"}

CV_VERSION_DIRS = {
    "CV1": "five-fold-cv",
    "CV2": "five-fold-cv2",
    "CV3": "five-fold-cv3",
    "CV4": "five-fold-cv4",
    "CV5": "five-fold-cv5",
    "CV6": "five-fold-cv6",
    "CV7": "five-fold-cv7",
}


def split_kind(split_dir: str) -> str:
    if split_dir.endswith("-as-test"):
        return "test"
    if split_dir.endswith("-as-training"):
        return "train"
    return "unknown"


def build_inventory(data_root: Path = DATA_ROOT, inspect_images: bool = True) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for path in sorted(data_root.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in IMAGE_EXTENSIONS:
            continue

        rel = path.relative_to(data_root)
        if len(rel.parts) < 5:
            continue

        version_dir, fold, split_dir, label = rel.parts[:4]
        row: dict[str, object] = {
            "version_dir": version_dir,
            "cv": _cv_from_version_dir(version_dir),
            "fold": fold,
            "split_dir": split_dir,
            "split": split_kind(split_dir),
            "label": label,
            "filename": path.name,
            "path": str(path),
            "extension": path.suffix.lower(),
            "bytes": path.stat().st_size,
            "width": pd.NA,
            "height": pd.NA,
            "mode": pd.NA,
            "readable": pd.NA,
            "error": "",
        }

        if inspect_images:
            try:
                with Image.open(path) as img:
                    row.update(width=img.width, height=img.height, mode=img.mode, readable=True)
            except (OSError, UnidentifiedImageError) as exc:
                row.update(readable=False, error=type(exc).__name__)

        rows.append(row)

    columns = [
        "cv",
        "version_dir",
        "fold",
        "split",
        "split_dir",
        "label",
        "filename",
        "extension",
        "bytes",
        "width",
        "height",
        "mode",
        "readable",
        "error",
        "path",
    ]
    return pd.DataFrame(rows, columns=columns)


def summarize_inventory(inventory: pd.DataFrame) -> pd.DataFrame:
    if inventory.empty:
        return pd.DataFrame(columns=["cv", "version_dir", "fold", "split", "split_dir", "label", "actual_count"])

    group_cols = ["cv", "version_dir", "fold", "split", "split_dir", "label"]
    return (
        inventory.groupby(group_cols, dropna=False)
        .size()
        .rename("actual_count")
        .reset_index()
        .sort_values(group_cols)
    )


def load_expected_counts(workbook_path: Path = SPLIT_WORKBOOK) -> pd.DataFrame:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[dict[str, object]] = []

    for cv in wb.sheetnames:
        ws = wb[cv]
        version_dir = CV_VERSION_DIRS.get(cv, cv)
        class_headers: list[str] | None = None
        current_fold: str | None = None

        for row in ws.iter_rows(values_only=True):
            values = list(row)
            if _is_class_header(values):
                class_headers = [str(v) for v in values[2:] if v not in (None, "total")]
                continue

            if not class_headers or not _is_split_row(values):
                continue

            if isinstance(values[0], str) and values[0].startswith("fold-"):
                current_fold = values[0]

            if current_fold is None:
                continue

            fold = current_fold
            split_dir = str(values[1])
            for label, expected in zip(class_headers, values[2:]):
                if expected is None:
                    continue
                rows.append(
                    {
                        "cv": cv,
                        "version_dir": version_dir,
                        "fold": fold,
                        "split": split_kind(split_dir),
                        "split_dir": split_dir,
                        "label": label,
                        "expected_count": int(expected),
                    }
                )

    columns = ["cv", "version_dir", "fold", "split", "split_dir", "label", "expected_count"]
    return pd.DataFrame(rows, columns=columns)


def compare_counts(actual: pd.DataFrame, expected: pd.DataFrame) -> pd.DataFrame:
    keys = ["cv", "version_dir", "fold", "split", "split_dir", "label"]
    comparison = expected.merge(actual, on=keys, how="outer")
    comparison["expected_count"] = pd.to_numeric(comparison["expected_count"], errors="coerce").fillna(0).astype(int)
    comparison["actual_count"] = pd.to_numeric(comparison["actual_count"], errors="coerce").fillna(0).astype(int)
    comparison["difference"] = comparison["actual_count"] - comparison["expected_count"]
    comparison["status"] = comparison["difference"].map(lambda x: "match" if x == 0 else "mismatch")
    return comparison.sort_values(keys).reset_index(drop=True)


def run_inspection(
    data_root: Path = DATA_ROOT,
    workbook_path: Path = SPLIT_WORKBOOK,
    output_dir: Path = METRICS_DIR,
    inspect_images: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    output_dir.mkdir(parents=True, exist_ok=True)

    inventory = build_inventory(data_root=data_root, inspect_images=inspect_images)
    actual_counts = summarize_inventory(inventory)
    expected_counts = load_expected_counts(workbook_path=workbook_path)
    comparison = compare_counts(actual_counts, expected_counts)

    inventory.to_csv(output_dir / "dataset_inventory.csv", index=False)
    actual_counts.to_csv(output_dir / "actual_counts.csv", index=False)
    expected_counts.to_csv(output_dir / "expected_counts.csv", index=False)
    comparison.to_csv(output_dir / "count_comparison.csv", index=False)
    return inventory, expected_counts, comparison


def _cv_from_version_dir(version_dir: str) -> str:
    for cv, dirname in CV_VERSION_DIRS.items():
        if dirname == version_dir:
            return cv
    return version_dir


def _is_class_header(values: list[object]) -> bool:
    if len(values) < 3:
        return False
    labels = {str(v) for v in values[2:] if v is not None}
    return "total" in labels and bool(labels & {"crater", "dark_dune", "edge", "streak", "bright_dune", "other"})


def _is_split_row(values: list[object]) -> bool:
    if len(values) < 2 or not isinstance(values[1], str):
        return False
    return split_kind(values[1]) in {"train", "test"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inspect HiRISE cross-validation image folders.")
    parser.add_argument("--data-root", type=Path, default=DATA_ROOT)
    parser.add_argument("--workbook", type=Path, default=SPLIT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=METRICS_DIR)
    parser.add_argument("--skip-image-open", action="store_true", help="Count paths without opening image files.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    inventory, _, comparison = run_inspection(
        data_root=args.data_root,
        workbook_path=args.workbook,
        output_dir=args.output_dir,
        inspect_images=not args.skip_image_open,
    )
    mismatches = int((comparison["status"] == "mismatch").sum())
    unreadable = 0 if inventory.empty or "readable" not in inventory else int((inventory["readable"] == False).sum())
    print(f"Images found: {len(inventory):,}")
    print(f"Count mismatches: {mismatches:,}")
    print(f"Unreadable images: {unreadable:,}")
    print(f"Saved CSV summaries to: {args.output_dir}")


if __name__ == "__main__":
    main()